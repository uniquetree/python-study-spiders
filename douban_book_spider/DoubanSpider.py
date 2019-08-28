#-*- coding: UTF-8 -*-
import random
import time
import urllib

import requests
import numpy as np
from bs4 import BeautifulSoup

from openpyxl import Workbook

userAgents=['Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36']

def random_create_headers():

    user_agent = userAgents[random.randint(1,5) % len(userAgents)]
    headers = {'User-Agent': user_agent,
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8'}
    return headers

# 豆瓣图书标签抓取
def book_tag_spider():
    tag_lists = []
    url = 'https://book.douban.com/tag/?view=type'

    session = requests.Session()
    req = session.get(url, headers=random_create_headers())

    bs = BeautifulSoup(req.text, 'html.parser')
    tables = bs.find_all('table', {'class': 'tagCol'})
    # 根据页面内容遍历豆瓣所有的书籍标签
    for table in tables:
        for book_tag in table.find_all('a'):
            tag_lists.append(book_tag.get_text())
    return tag_lists


# 根据图书标签爬取豆瓣图书信息
def book_info_spider(book_tag):
    page_num = 0;
    book_infos = []
    try_times = 0

    while 1:
        time.sleep(np.random.rand() * 1)
        try_times += 1;

        url = 'https://book.douban.com/tag/' + urllib.parse.quote(book_tag) + '?start= ' + str(page_num * 20) + '&type=T'

        session = requests.Session()
        req = session.get(url, headers=random_create_headers())
        soup = BeautifulSoup(req.text, 'html.parser')
        list_soup = soup.find('ul', {'class': 'subject-list'}).findAll('li', {'class': 'subject-item'})

        if list_soup is None or len(list_soup) <= 1 or try_times > 20:
            break

        for book_item in list_soup:
            book_info = book_item.find('div', {'class': 'info'})

            title = book_info.findAll('a')[0]['title']
            book_detail_info = book_info.find('div', {'class': 'pub'}).get_text().strip().split('/')
            book_url = book_info.findAll('a')[0].get('href')

            # 作者/译者
            try:
                author_info = ''.join(book_detail_info[0:-3])
            except:
                author_info = '暂无'
            # 出版信息
            try:
                pub_info = ''.join(book_detail_info[-3:])
            except:
                pub_info = '暂无'
            # 评分
            try:
                rating = book_info.find('span', {'class': 'rating_nums'}).get_text().strip()
            except:
                rating = '0.0'
            # 评价人数
            try:
                people_num = book_info.find('span', {'class': 'pl'}).get_text().strip().strip('()人评价')
            except:
                people_num = '0'

            # 保存书籍信息
            book_infos.append([title, rating, people_num, author_info, pub_info])

        page_num += 1

    # print('Get book infos success! [%s]' % book_infos)
    return book_infos


# 将爬取到到图书信息保存到csv文件，文件命名为'图书标签名.csv'
def print_book_lists_excel(book_info_lists):
    wb = Workbook()
    ws = []
    for i in range(len(book_info_lists)):
        ws.append(wb.create_sheet(title=book_info_lists[i].get('book_tag_name')))

        ws[i].append(['序号', '书名', '评分', '评价人数', '作者', '出版信息'])
        count = 1
        for bl in book_info_lists[i].get('book_infos'):
            ws[i].append([count, bl[0], float(bl[1]), int(bl[2]), bl[3], bl[4]])
            count += 1
    wb.save('book_list.xlsx')


if __name__ == '__main__':
    # 图书列表对象，('图书标签': '图书标签下到图书信息')
    book_tag_lists = []
    for book_tag in book_tag_spider():
        book_tag_list = {}
        if book_tag == '编程':
            book_tag_list['book_infos'] = book_info_spider(book_tag)
            book_tag_list['book_tag_name'] = book_tag
            book_tag_lists.append(book_tag_list)
            break
    print_book_lists_excel(book_tag_lists)