#-*- coding: UTF-8 -*-
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

userAgents=['Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36']

# 模拟请求头
def random_create_headers():
    user_agent = userAgents[random.randint(1,5) % len(userAgents)]
    headers = {'User-Agent': user_agent,
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8'}
    return headers

# 豆瓣图书标签抓取
def book_tag_spider():
    tag_lists = []
    url = 'https://book.douban.com/tag/?view=type'
    session = requests.Session()
    req = session.get(url, headers=random_create_headers())
    bs = BeautifulSoup(req.text, 'html.parser')
    tables = bs.find_all('table', {'class': 'tagCol'})
    # 根据页面内容遍历豆瓣所有的书籍标签
    for table in tables:
        for book_tag in table.find_all('a'):
            tag_lists.append(book_tag.get_text())
    return tag_lists

# 将爬取到到图书信息保存到csv文件，文件命名为'图书标签名.csv'
def print_book_lists_excel(book_tags):
    wb = Workbook()
    ws = []
    count = 1
    ws.append(wb.create_sheet(title='图书标签'))
    ws[0].append(['序号', '图书标签'])
    for i in range(len(book_tags)):
        ws[0].append([count, book_tags[i]])
        count += 1
    wb.save('图书标签.xlsx')

if __name__ == '__main__':
    print_book_lists_excel(book_tag_spider())